import random
import xlsxwriter
from openpyxl import load_workbook
from math import floor


# def avg_bp():
#     workbook = load_workbook(filename="blood_pressure.xlsx")
#     sys = 0
#     dia = 0
#     sheet1 = workbook['Sheet1']
#     for i in range(2, 7):
#         for j in range(1, 8):
#             y = (sheet1[i][j].value).split('/')
#             sys += int(y[0])
#             dia += int(y[1])
#         print("Average Blood Pressure is:" +
#               str(floor(sys/7)) + "/" + str(floor(dia/7)))
#         sys = 0
#         dia = 0


# def avg_calories():
#     workbook = load_workbook(filename="calories.xlsx")
#     cal_high = ["Great Job you are burning great amounts of calories", "Amazing... Keep up the good workout",
#                 "Keep burning more calories", "Your metabolic rate is high !!!1", "You definitely perform a lot of cardiovascular excercises..."]
#     cal_low = ["You should workout more.", "Try to walk everyday", "You might have a sedentary lifestyle try working out",
#                "You can do better...", "Push yourself a little more to achieve wonders"]
#     cal_mid = ["You are doing great", "DON'T stop keep burning",
#                "You are on the right track to staying fit", "Keep it up and Never give up"]
#     cal = 0
#     sheet1 = workbook['Sheet1']
#     for i in range(2, 7):
#         for j in range(1, 8):
#             cal += (sheet1[i][j].value)
#         print("Average Calaories burned are :" + str(floor(cal/7)), end=" ")
#         if floor(cal/7) > 2100:
#             print(random.choice(cal_high))
#         elif(floor(cal/7) < 1900):
#             print(random.choice(cal_low))
#         else:
#             print(random.choice(cal_mid))
#         cal = 0


# def avg_sugar():
#     workbook = load_workbook(filename="sugar_levels.xlsx")
#     sug = 0
#     sheet1 = workbook['Sheet1']
#     for i in range(2, 7):
#         for j in range(1, 8):
#             sug += (sheet1[i][j].value)
#         print("Average Sugar levels observed are :" + str(floor(sug/7)))
#         sug = 0


# # workbook = xlsxwriter.Workbook('blood_pressure.xlsx')
# # workbook1 = xlsxwriter.Workbook('calories.xlsx')
# # workbook2 = xlsxwriter.Workbook('sugar_levels.xlsx')
# # worksheet = workbook.add_worksheet()
# # worksheet1 = workbook1.add_worksheet()
# # worksheet2 = workbook2.add_worksheet()
# # bp = ["high", "normal", "low"]
# # for j in range(1, 6):
# #     for i in range(1, 8):
# #         x = random.randint(0, 2)
# #         y = bp[x]

# #         if(y == "high"):
# #             worksheet.write(j, i, str(random.randint(83, 95)) + "/" +
# #                             str(random.randint(124, 180)))
# #             worksheet2.write(j, i, random.randint(200, 300))

# #         elif(y == "normal"):
# #             worksheet.write(j, i, str(random.randint(78, 83)) + "/" +
# #                             str(random.randint(118, 124)))
# #             worksheet2.write(j, i, random.randint(140, 199))

# #         else:
# #             worksheet.write(j, i, str(random.randint(78, 83)) + "/" +
# #                             str(random.randint(118, 124)))
# #             worksheet2.write(j, i, random.randint(80, 139))

# #         worksheet1.write(j, i, random.randint(1600, 2300))
# # workbook.close()
# # workbook1.close()
# # workbook2.close()

# def daily_bp(a):
# workbook = load_workbook(filename="blood_pressure.xlsx")
# sheet1 = workbook['Sheet1']
# for i in range(1, 8):
#     print(sheet1[a+1][i].value, end=" ")
# print()


# def daily_cal(a):
#     workbook = load_workbook(filename="calories.xlsx")
#     sheet1 = workbook['Sheet1']
#     for i in range(1, 8):
#         print(sheet1[a+1][i].value, end=" ")
#     print()


# def daily_sugar(a):
#     workbook = load_workbook(filename="sugar_levels.xlsx")
#     sheet1 = workbook['Sheet1']
#     for i in range(1, 8):
#         print(sheet1[a+1][i].value, end=" ")
#     print()


# # avg_bp()
# avg_calories()
# # avg_sugar()

# # daily_bp(3)
# # daily_cal(4)
# daily_sugar(3)

# workbook1 = load_workbook(filename="blood_pressure.xlsx")
# worksheet1 = workbook1['Sheet1']
# for i in range(1, 6):
#     if worksheet1.cell(i, 1).value == None:
#         worksheet1.cell(i, 1).value = "mohit"
#         break
# # await ctx.channel.send(">"+ctx.author.name+"is being added")
# workbook1.close()
import prettytable
from prettytable import PrettyTable

l = [["Hassan", 21, "LUMS"], ["Ali", 22, "FAST"], ["Ahmed", 23, "UET"]]

table = PrettyTable(['Name', 'Age', 'University'])

for rec in l:
    table.add_row(rec)

print(table)
