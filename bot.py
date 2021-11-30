# python libraries

from inspect import _empty
import discord
from discord import colour
from discord import channel
from openpyxl import descriptors
import xlsxwriter
from discord.ext import commands
import discord.utils
from discord.utils import *
import random
from openpyxl import *
from math import floor
import requests
from table2ascii import table2ascii as t2a

bot = commands.Bot(command_prefix="-", help_command=None)

# bot = discord.Client()


# Dataset

# @bot.event
# async def on_message(message):
#     workbook = xlsxwriter.Workbook('blood_pressure.xlsx')
#     workbook1 = xlsxwriter.Workbook('calories.xlsx')
#     workbook2 = xlsxwriter.Workbook('sugar_levels.xlsx')
#     worksheet = workbook.add_worksheet()
#     worksheet1 = workbook1.add_worksheet()
#     worksheet2 = workbook2.add_worksheet()
#     bp = ["high", "normal", "low"]
#     for j in range(1, 6):
#         for i in range(1, 8):
#             x = random.randint(0, 2)
#             y = bp[x]

#             if(y == "high"):
#                 worksheet.write(j, i, str(random.randint(124, 180))+"/" +
#                                 str(random.randint(83, 95)))
#                 worksheet2.write(j, i, random.randint(200, 300))

#             elif(y == "normal"):
#                 worksheet.write(j, i, str(random.randint(118, 124))+"/" +
#                                 str(random.randint(78, 83)))
#                 worksheet2.write(j, i, random.randint(140, 199))

#             else:
#                 worksheet.write(j, i,  str(random.randint(118, 124))+"/" +
#                                 str(random.randint(78, 83)))
#                 worksheet2.write(j, i, random.randint(80, 139))

#             worksheet1.write(j, i, random.randint(1600, 2300))
#     workbook.close()
#     workbook1.close()
#     workbook2.close()
#     await bot.process_commands(message)


l = []


@bot.command(name="help")
async def help(ctx):
    mbed = discord.Embed(
        colour=(discord.Colour.magenta()),
        title='I am Rob, your personal Healthcare Assistant.',
        description="Hello... Glad to see you here.\n\nCurious to know your health parameters such as blood pressure, sugar levels and calories burned in a day?\n\n Don't worry I will provide all the data you need along with a tini-tiny bit of friendly advice.\n "+"\_"*74
    )
    mbed.add_field(name="Average Blood Pressure",
                   value="I will display your average Blood Pressure levels during the whole week.\n\n Use '-avg_bp'\n command to know more !\n "+"\_"*23, inline=True)
    mbed.add_field(name="Average Sugar levels",
                   value="I will display your average sugar levels during the whole week.\n\n Use '-avg_sugar'\n command to know more !\n"+"\_"*23, inline=True)
    mbed.add_field(name="Average Calories burned",
                   value="I will display your average calories burned during the whole week.\n\n Use '-avg_calories'\n command to know more !\n"+"\_"*23, inline=True)
    # mbed.add_field()
    mbed.add_field(name="Daily Blood Pressure",
                   value="I will display your Blood Pressure levels for the past whole week.\n\n Use '-daily_bp'\n command to know more !\n "+"\_"*23, inline=True)
    mbed.add_field(name="Daily Sugar levels",
                   value="I will display your Sugar levels for the past whole week.\n\n Use '-daily_sugar'\n command to know more !\n "+"\_"*23, inline=True)
    mbed.add_field(name="Daily Calories burned",
                   value="I will display your Calories burned for the past whole week.\n\n Use '-daily_calories'\n command to know more !\n "+"\_"*23, inline=True)
    mbed.add_field(name="Motivation", value="Feeling low and need some motivation ?? I got it covered for you.\n Just type '-motivate' to get a inspirational /motivational quote from the great minds.\n"+"\_"*74, inline=False)
    mbed.set_footer(
        text="All the data displayed is calculated from the sample dataset stored in the local machine.")
    await ctx.channel.send(embed=mbed)


@bot.command("make")
async def make(ctx):
    workbook1 = load_workbook(filename="blood_pressure.xlsx")
    workbook2 = load_workbook(filename="sugar_levels.xlsx")
    workbook3 = load_workbook(filename="calories.xlsx")
    sheet1 = workbook1.active
    sheet2 = workbook2.active
    sheet3 = workbook3.active
    y = ctx.author.name+ctx.author.discriminator
    if(y in l):
        await ctx.channel.send("> You are already registered in the database")
    # await ctx.channel.send(">"+ctx.author.name+"is being added")
    else:
        l.append(y)
        for i in range(2, 7):
            if(sheet1.cell(i, 1).value == None):
                sheet1.cell(i, 1).value = ctx.author.name + \
                    ctx.author.discriminator
                sheet2.cell(i, 1).value = ctx.author.name + \
                    ctx.author.discriminator
                sheet3.cell(i, 1).value = ctx.author.name + \
                    ctx.author.discriminator
                await ctx.send("> The data for"+ctx.author.mention+" has been crated successfully")
                break
    workbook1.save(filename="blood_pressure.xlsx")
    workbook2.save(filename="sugar_levels.xlsx")
    workbook3.save(filename="calories.xlsx")


@bot.command(name="avg_bp")
async def bp(ctx):
    # await ctx.channel.send(ctx.author.id)
    workbook = load_workbook(filename="blood_pressure.xlsx")
    sheet1 = workbook.active
    sys = 0
    dia = 0
    sheet1 = workbook['Sheet1']
    bp_high = ["Take some rest.",
               "Lie down and take deep breaths.",
               "Try to stay calm and relax."]
    bp_normal = ["Have a great day ahead.",
                 "Nothing to worry."]
    bp_low = ["Drink plenty of water.",
              "Increase salt intake.",
              "Take rest!"]
    y = ctx.author.name+ctx.author.discriminator
    if(y in l):
        for i in range(2, 7):
            if(sheet1.cell(i, 1).value == ctx.author.name+ctx.author.discriminator):
                for j in range(1, 8):
                    response = ""
                    res = ""
                    y = (sheet1[i][j].value).split('/')
                    sys += int(y[0])
                    dia += int(y[1])
                if(floor(int(sys/7)) > 120):
                    response += random.choice(bp_high)
                    res += "HIGH Blood Pressure"
                elif(floor(int(sys/7)) >= 80 and floor(int(sys/7)) <= 120):
                    response += random.choice(bp_normal)
                    res += "NORMAL Blood Pressure"
                else:
                    response += random.choice(bp_low)
                    res += "LOW Blood Pressure"
                mbed = discord.Embed(
                    colour=(discord.Colour.magenta()),
                    title='Average Blood Pressure over a span of 7 days',
                    description=ctx.author.mention+" Your average Blood Pressure is: " +
                    str(floor(sys/7)) + "/" + str(floor(dia/7)) + "mmHg")
                mbed.add_field(name="Category:", value=res, inline=True)
                mbed.add_field(name="My Advice:",
                               value="Beep Beep... "+response, inline=True)
                mbed.set_footer(
                    text="The average has been calculated from the sample dataset stored in the local machine")
                await ctx.channel.send(embed=mbed)
                sys = 0
                dia = 0
    else:
        await ctx.channel.send(
            "> Could not find you in the database please use '-make' command to register you into the database")


@bot.command(name="avg_sugar")
async def sugar(ctx):
    y = ctx.author.name+ctx.author.discriminator
    if(y in l):
        workbook = load_workbook(filename="sugar_levels.xlsx")
        sug = 0
        sheet1 = workbook['Sheet1']
        sug_high = ["Drink More Water.",
                    "Eat green & healthy food.",
                    "Exercise and Eat healthy."]
        sug_normal = ["Have a great day Ahead.",
                      "Keep Eating healthy & Exercise."]
        sug_low = ["Eat more carbohydrates.",
                   "Try fruit juice or Biscuit.",
                   "Try eating few sweets because you sugar levels are below the normal level"]
        for i in range(2, 7):
            if(sheet1.cell(i, 1).value == ctx.author.name+ctx.author.discriminator):
                for j in range(1, 8):
                    response = ""
                    res = ""
                    sug += (sheet1[i][j].value)
                    if(floor(sug/7) > 200):
                        response += random.choice(sug_high)
                        res += "HIGH Sugar levels !!!!"
                    elif(floor(sug/7) >= 140 and floor(sug/7) <= 200):
                        response += random.choice(sug_normal)
                        res += "NORMAL Sugar levels"
                    else:
                        response += random.choice(sug_low)
                        res += "LOW Sugar levels"
                mbed = discord.Embed(
                    colour=(discord.Colour.magenta()),
                    title='Average sugar levels observed over a span of 7 days',
                    description=f'''{ctx.author.mention} Your average Sugar Levels observed are : {str(floor(sug/7))} mg/L'''
                    # ctx.author.mention +
                    # "Your average Sugar Levels observed are : " +
                )
                mbed.add_field(name="Category:", value=res, inline=True)
                mbed.add_field(name="My Advice:",
                               value="Beep Beep Boop.. "+response, inline=True)
                mbed.set_footer(
                    text="The average has been calculated from the sample dataset stored in the local machine")
                await ctx.channel.send(embed=mbed)
                sug = 0
    else:
        await ctx.channel.send(
            "> Could not find you in the database please use '-make' command to register you into the database")


@bot.command(name="avg_calories")
async def calories(ctx):
    y = ctx.author.name+ctx.author.discriminator
    if(y in l):
        workbook = load_workbook(filename="calories.xlsx")
        cal_high = ["Great Job you are burning great amounts of calories",
                    "Amazing... Keep up the good workout",
                    "Keep burning more calories",
                    "You definitely perform a lot of cardiovascular excercises..."]
        cal_low = ["You should workout more.",
                   "Try to walk everyday to burn more calories",
                   "You might have a sedentary lifestyle try working out",
                   "Try burning a few more calories everyday.",
                   "Push yourself a little more to achieve wonders"]
        cal_mid = ["Enjoy and keep doing the excercises you do.",
                   "DON'T stop you are maintaing a great work-life balance",
                   "Follow your daily routine... You are FIT",
                   "Keep it up and Never give up"]
        cal = 0
        sheet1 = workbook['Sheet1']
        for i in range(2, 7):
            if(sheet1.cell(i, 1).value == ctx.author.name+ctx.author.discriminator):
                for j in range(1, 8):
                    cal += (sheet1[i][j].value)
                response = ""
                if floor(cal/7) > 2100:
                    response += (random.choice(cal_high))
                elif(floor(cal/7) < 1900):
                    response += (random.choice(cal_low))
                else:
                    response += (random.choice(cal_mid))
                mbed = discord.Embed(
                    colour=(discord.Colour.magenta()),
                    title='Average calories burned over a span of 7 days',
                    description="Average Calories Burned for "+ctx.author.mention+" are : " +
                    str(floor(cal/7))+"Kcal"
                )
                mbed.add_field(name="My Advice:",
                               value="Beep Boop.. "+response, inline=False)
                mbed.set_footer(
                    text="The average has been calculated from the sample dataset stored in the local machine")
                await ctx.channel.send(embed=mbed)
                cal = 0
    else:
        await ctx.channel.send(
            "> Could not find you in the database please use '-make' command to register you into the database")


@bot.command(name="daily_bp")
async def bp(ctx):
    y = ctx.author.name+ctx.author.discriminator
    if(y in l):
        workbook = load_workbook(filename="blood_pressure.xlsx")
        sheet1 = workbook['Sheet1']
        for i in range(2, 7):
            if(sheet1.cell(i, 1).value == ctx.author.name+ctx.author.discriminator):
                response = ""
                for j in range(1, 8):
                    response += "Day "+str(j)+": " + \
                        sheet1[i][j].value+" mmHg\n"
                    y = (sheet1[i][j].value).split('/')
                    if(int(y[0]) > 120):
                        response += "Your BP is High\n\n"
                    elif(int(y[0]) < 110):
                        response += "Your BP is Low\n\n"
                    else:
                        response += "Your BP is Normal \n\n"
                mbed = discord.Embed(
                    colour=(discord.Colour.magenta()),
                    title='Daily data of Blood Pressure',
                    description="Here daily data of Blood pressure is displayed: "
                )
                mbed.add_field(name="Data",
                               value=response)
                mbed.set_footer(
                    text="The daily has been displayed from the sample dataset stored in the local machine")
                await ctx.channel.send(embed=mbed)
                # await ctx.channel.send("Day "+str(j)+": "+sheet1[i][j].value)
    else:
        await ctx.channel.send(
            "> Could not find you in the database please use '-make' command to register you into the database")


@bot.command(name="daily_sugar")
async def bp(ctx):
    y = ctx.author.name+ctx.author.discriminator
    if(y in l):
        workbook = load_workbook(filename="sugar_levels.xlsx")
        sheet1 = workbook['Sheet1']
        for i in range(2, 7):
            if(sheet1.cell(i, 1).value == ctx.author.name+ctx.author.discriminator):
                response = ""
                for j in range(1, 8):
                    response += "Day "+str(j)+": " + \
                        str(sheet1[i][j].value)+" mg/L\n"
                    if(sheet1[i][j].value > 200):
                        response += "Your SUgar levels are High\n\n"
                    elif(sheet1[i][j].value >= 140 and sheet1[i][j].value <= 200):
                        response += "Your Sugar levels are Normal\n\n"
                    else:
                        response += "Your Sugar levels are Low\n\n"
                mbed = discord.Embed(
                    colour=(discord.Colour.magenta()),
                    title='Daily data of Sugar Levels',
                    description="Here data of daily sugar levels on a day is displayed: "
                )
                mbed.add_field(name="Data",
                               value=response)
                mbed.set_footer(
                    text="The daily has been displayed from the sample dataset stored in the local machine")
                await ctx.channel.send(embed=mbed)
    else:
        await ctx.channel.send(
            "> Could not find you in the database please use '-make' command to register you into the database")


@bot.command(name="daily_calories")
async def bp(ctx):
    y = ctx.author.name+ctx.author.discriminator
    if(y in l):
        workbook = load_workbook(filename="calories.xlsx")
        sheet1 = workbook['Sheet1']
        for i in range(2, 7):
            if(sheet1.cell(i, 1).value == ctx.author.name+ctx.author.discriminator):
                response = ""
                for j in range(1, 8):
                    response += "Day "+str(j)+": " + \
                        str(sheet1[i][j].value)+" Kcal\n"
                mbed = discord.Embed(
                    colour=(discord.Colour.magenta()),
                    title='Daily data of Calories Burned',
                    description="Here data of daily calories burned is shown: "
                )
                mbed.add_field(name="Data",
                               value=response)
                mbed.set_footer(
                    text="The daily has been displayed from the sample dataset stored in the local machine")
                await ctx.channel.send(embed=mbed)
    else:
        await ctx.channel.send(
            "> Could not find you in the database please use '-make' command to register you into the database")


@bot.command(name="motivate")
async def motivate(ctx):
    response = requests.get("https://zenquotes.io/api/random")
    data = json.loads(response.text)
    await ctx.channel.send(data[0]['q']+" ~~ "+data[0]['a'])


@bot.event
async def on_ready():
    print('I am {0.user}'.format(bot))


bot.run(TOKEN)
